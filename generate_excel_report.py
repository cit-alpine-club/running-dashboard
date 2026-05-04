"""
results_all_members.csv を整形し、メンバーごとのシートと
まとめシート（グラフ付き）を含む Excel ファイルを生成する。

使い方:
  .venv/Scripts/python.exe generate_excel_report.py
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime


def normalize_time(time_str):
    """時間表記を統一する（末尾の :00 を削除）"""
    if pd.isna(time_str) or time_str == '--':
        return '--'
    time_str = str(time_str).strip()
    if time_str.endswith(':00'):
        return time_str[:-3]
    return time_str


def generate_excel_report(csv_path, output_path):
    """CSV を読み込み、Excel レポートを生成する"""
    
    # CSV を読み込む
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    
    # ファイル名列を削除
    df = df.drop(columns=['ファイル名'], errors='ignore')
    
    # 時間列を正規化
    if '時間' in df.columns:
        df['時間'] = df['時間'].apply(normalize_time)
    
    # メンバーごとにグループ化
    grouped = df.groupby('投稿者')
    
    # Excel ワークブックを作成
    wb = Workbook()
    wb.remove(wb.active)  # 空のシートを削除
    
    # メンバーごとのシートを作成
    member_summary = {}
    
    for member_name, member_df in grouped:
        # 距離を数値に変換
        member_df = member_df.copy()
        member_df['距離(km)'] = pd.to_numeric(member_df['距離(km)'], errors='coerce')
        
        # 距離の合計値を計算
        total_distance = member_df['距離(km)'].sum()
        member_summary[member_name] = {
            'total_distance': total_distance,
            'count': len(member_df),
            'data': member_df
        }
        
        # シートを作成
        ws = wb.create_sheet(title=member_name[:31])  # シート名は31文字まで
        
        # ヘッダー行を書き込む
        headers = member_df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # データを書き込む
        for row_idx, row in enumerate(member_df.values, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 合計行を追加
        total_row = len(member_df) + 2
        ws.cell(row=total_row, column=1).value = '合計'
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # 距離(km) 列の合計値を追加
        distance_col = headers.index('距離(km)') + 1 if '距離(km)' in headers else None
        if distance_col:
            cell = ws.cell(row=total_row, column=distance_col)
            cell.value = total_distance
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # 列幅を自動調整
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = max(len(header), 12)
    
    # まとめシートを作成
    summary_ws = wb.create_sheet('まとめ', 0)
    
    # タイトル
    summary_ws['A1'] = 'メンバー別ランニング記録 - サマリー'
    summary_ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    summary_ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_ws.merge_cells('A1:D1')
    
    # サマリー表を作成
    summary_ws['A3'] = 'メンバー名'
    summary_ws['B3'] = '距離合計(km)'
    summary_ws['C3'] = '走行数'
    
    for col in ['A', 'B', 'C']:
        cell = summary_ws[f'{col}3']
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    row = 4
    colors = ['FFE699', 'B4C7E7', 'C6E0B4', 'F8CBAD', 'E2EFDA', 'FCE4D6', 'EDEDED']
    color_idx = 0
    
    for member_name in sorted(member_summary.keys()):
        summary = member_summary[member_name]
        summary_ws[f'A{row}'] = member_name
        summary_ws[f'B{row}'] = summary['total_distance']
        summary_ws[f'C{row}'] = summary['count']
        
        color = colors[color_idx % len(colors)]
        for col in ['A', 'B', 'C']:
            cell = summary_ws[f'{col}{row}']
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        color_idx += 1
        row += 1
    
    # 列幅を設定
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15
    summary_ws.column_dimensions['C'].width = 12
    
    # 棒グラフ（メンバーごとの距離合計）を作成
    bar_chart = BarChart()
    bar_chart.type = 'col'
    bar_chart.title = 'メンバー別 距離合計（km）'
    bar_chart.y_axis.title = '距離（km）'
    bar_chart.x_axis.title = 'メンバー'
    
    data_rows = len(member_summary)
    values = Reference(summary_ws, min_col=2, min_row=3, max_row=3 + data_rows)
    cats = Reference(summary_ws, min_col=1, min_row=4, max_row=3 + data_rows)
    bar_chart.add_data(values, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.height = 10
    bar_chart.width = 20
    
    summary_ws.add_chart(bar_chart, 'A10')
    
    # Excel ファイルを保存
    wb.save(output_path)
    print(f'✓ Excel レポート生成: {output_path}')


if __name__ == '__main__':
    csv_path = Path('./screenshots/results_all_members.csv')
    output_path = Path('./screenshots/results_all_members.xlsx')
    
    if not csv_path.exists():
        print(f'エラー: {csv_path} が見つかりません')
        exit(1)
    
    generate_excel_report(csv_path, output_path)
